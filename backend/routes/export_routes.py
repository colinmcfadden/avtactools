from flask import Blueprint, request, send_file, jsonify, current_app
import zipfile
import os
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO

from export_service import generate_custom_package

export_bp = Blueprint('export_bp', __name__)

@export_bp.route('/api/generate-excel', methods=['POST'])
def generate_excel():
    # 1. Get Data and Image from the request
    # (Assuming you send data as form-data: 'json_data' and 'map_image')
    data = request.form 
    image_file = request.files['map_image']

    # 2. Load the Template
    template_path = 'lz_template.xlsx'
    wb = load_workbook(template_path)
    ws = wb.active  # Get the first sheet

    # 3. Fill in the Text Data (Mapping specific cells)
    ws['A1'] = data.get('lz_label')
    ws['B1'] = data.get('lz_name')

    ws['D2'] = data.get('objective')
    ws['D3'] = data.get('mgrs_grid')
    ws['D4'] = data.get('lat_long')
    ws['D5'] = data.get('elevation')
    
    # Bottom Data Block
    ws['A23'] = data.get('call_sign')
    ws['C23'] = data.get('freq')
    ws['D23'] = data.get('formation')
    ws['E23'] = data.get('land_dir')
    ws['H23'] = data.get('go_around')
    ws['J23'] = data.get('takeoff_dir')
    ws['J25'] = data.get('formation')
    ws['F24'] = data.get('door')
    ws['F25'] = data.get('load')
    ws['D25'] = data.get('weapons_status')
    ws['A26'] = data.get('remarks')

    ws['A28'] = data.get('lz_label')
    ws['B28'] = data.get('lz_name')

    # 4. Insert the Map Image
    # Need to process the image stream to make it Excel-compatible
    #img_stream = BytesIO(image_file.read())
    img = ExcelImage(image_file)
    
    # Resize image to fit the cell
    img.width = 663
    img.height = 555
    
    # Anchor the image to a specific cell (Top-Left corner of the image)
    ws.add_image(img, 'A6')

    # 5. Save to Memory and Return
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    return send_file(
        output_stream,
        as_attachment=True,
        download_name=f"{data.get('lz_label', 'LZ')}_{data.get('lz_name', 'CARD')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@export_bp.route('/api/export-package', methods=['POST'])
def export_package():
    data = request.json
    
    # CHANGED: We now expect a dictionary like { "blue": [...], "red": [...] }
    # instead of a flat list.
    bounds_dict = data.get('bounds') 
    
    if not bounds_dict:
        return jsonify({"error": "No bounds provided"}), 400

    try:
        # 1. Generate the custom files (ForeFlight Tiff + Visual Jpg)
        # This function returns a list of file paths that were created on disk
        file_paths = generate_custom_package(bounds_dict)
        
        if not file_paths:
            return jsonify({"error": "No valid bounds received or file generation failed"}), 400

        # 2. Create a Zip file in memory
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for f_path in file_paths:
                # Add the file to the zip with its generated name
                zipf.write(f_path, arcname=os.path.basename(f_path))
                
                # Clean up the temp file from disk immediately after adding to zip
                try:
                    os.remove(f_path)
                except Exception as cleanup_error:
                    current_app.logger.warning(f"Warning: Could not cleanup temp file {f_path}: {cleanup_error}")
            
            # Optional: Add metadata
            zipf.writestr("metadata.txt", f"Export Configuration: {bounds_dict}")

        # 3. Send back to client
        memory_file.seek(0)

        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='Mission_Package.zip'
        )

    except Exception as e:
        current_app.logger.error(f"Export failed: {str(e)}")
        # Clean up any files that might have been created before the error
        # (This is a safety check)
        if 'file_paths' in locals():
            for f in file_paths:
                if os.path.exists(f):
                    try: os.remove(f)
                    except: pass
                    
        return jsonify({"error": str(e)}), 500